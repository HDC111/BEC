import os
import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import yaml

def run_matcher(config):

    # --- Paths from config ---
    processed_dir = os.path.join(config["paths"]["current_case"], "processed")
    output_access_file = os.path.join(processed_dir, "output_accessed.xlsx")
    output_matched_file = os.path.join(processed_dir, "matched_rows_from_suspicious_folders.xlsx")
    output_marked_file = os.path.join(processed_dir, "output_accessed_marked.xlsx")
    # --- Load main access log ---
    output_df = pd.read_excel(output_access_file)
    output_df.columns = output_df.columns.str.encode('ascii', 'ignore').str.decode('ascii')
    output_df.columns = output_df.columns.str.strip().str.upper()
    output_df["SUSPICIOUS"] = "no"

    matched_rows = []

    # --- Step 1: Find suspicious folders ---
    suspicious_roots = []
    for root, dirs, files in os.walk(processed_dir):
        for d in dirs:
            if d.lower().startswith("suspicious"):
                suspicious_roots.append(os.path.join(root, d))

    # --- Step 2: Gather suspicious .xlsx files ---
    suspicious_files = []
    for s_root in suspicious_roots:
        for root, _, files in os.walk(s_root):
            for f in files:
                if f.endswith(".xlsx"):
                    suspicious_files.append(os.path.join(root, f))

    print(f"📁 Found {len(suspicious_files)} suspicious .xlsx file(s).")

    # --- Step 3: Match ---
    for file_path in tqdm(suspicious_files, desc="📊 Matching rows"):
        try:
            df = pd.read_excel(file_path)
            df.columns = df.columns.str.encode('ascii', 'ignore').str.decode('ascii')
            df.columns = df.columns.str.strip().str.upper()

            required_keys = {"CREATIONTIME", "SESSIONID"}
            available_keys = required_keys.intersection(df.columns).intersection(output_df.columns)

            if not available_keys:
                print(f"⚠️ Skipping {file_path} — no shared key.")
                continue

            mask = pd.Series(False, index=output_df.index)
            for key in available_keys:
                if key == "CREATIONTIME":
                    df[key] = pd.to_datetime(df[key], errors="coerce").dt.strftime("%Y-%m-%d %H:%M:%S")
                    output_df[key] = pd.to_datetime(output_df[key], errors="coerce").dt.strftime("%Y-%m-%d %H:%M:%S")
                else:
                    df[key] = df[key].astype(str).str.strip().str.lower()
                    output_df[key] = output_df[key].astype(str).str.strip().str.lower()
                
                mask |= output_df[key].isin(df[key].dropna())

            output_df.loc[mask, "SUSPICIOUS"] = "yes"
            matched = output_df[mask]

            for _, m in matched.iterrows():
                matched_rows.append({
                    "Matched From": file_path,
                    **m.to_dict()
                })

        except Exception as e:
            print(f"❌ Error processing {file_path}: {e}")

    # --- Step 4: Save matched suspicious rows ---
    if matched_rows:
        df_result = pd.DataFrame(matched_rows)
        if "CREATIONTIME" in df_result.columns:
            df_result["CREATIONTIME"] = pd.to_datetime(df_result["CREATIONTIME"], errors='coerce')
            df_result = df_result.sort_values(by="CREATIONTIME")
        df_result.to_excel(output_matched_file, index=False)
        print(f"✅ Suspicious records saved to {output_matched_file}")
    else:
        print("⚠️ No suspicious records found.")

    # --- Step 5: Save full output with highlights ---
    output_df.to_excel(output_marked_file, index=False)
    wb = load_workbook(output_marked_file)
    ws = wb.active

    suspicious_col = None
    for i, col in enumerate(ws[1], start=1):
        if col.value and str(col.value).strip().upper() == "SUSPICIOUS":
            suspicious_col = i
            break

    if suspicious_col:
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for row in ws.iter_rows(min_row=2):
            if row[suspicious_col - 1].value == "yes":
                for cell in row:
                    cell.fill = fill

    wb.save(output_marked_file)
    print(f"✅ Full output with highlights saved to {output_marked_file}")
